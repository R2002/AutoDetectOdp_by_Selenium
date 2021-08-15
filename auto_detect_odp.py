#---------------------------------------#
# モジュール読み込み
#---------------------------------------#
import pandas as pd
from datetime import datetime
import re, time, os, tempfile
from ctypes import windll
import openpyxl, shutil
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.common import exceptions
from selenium.webdriver.common.action_chains import ActionChains
import Levenshtein as LS
# 自作関数
from src.beep import beep # beep(音の高さ, 長さ)

#---------------------------------------#
# 設定
#---------------------------------------#
# 待機時間
setting = {
    'time_wait':  5, # ファイル間の読み込み待機時間（読み込みが上手く行かない場合再設定してください）
    'time_wait_q':3, # 表示遅延
    'time_wait_file':10, # ファイルダウンロード用待機時間
    'count_retry':10, # 再試行回数
    'refusal_accuracy':0.90, # レーベンシュタイン距離・正答率許容度（文字数に対し何割までミスを許すか）
}

# データ設定
df_dict = {} # 各特許情報格納
dict_download = {} # ダウンロードデータ整合性確認
# 基礎情報の取得用DataFrame
df_all = pd.DataFrame(columns=['C.Code', '最新情報', '最終更新日', '出願番号', '出願日', '公開番号', '登録番号'])
# 各国別データcolumns
list_data_columns = ['Date', 'Reference', 'FileName', 'Key', 'Url']
# ダウンロードデータのリスト
list_download_files = []

# ブラウザウィンドウサイズ読み込み
window_w = windll.user32.GetSystemMetrics(0)  # 横幅
window_h = windll.user32.GetSystemMetrics(1)  # 縦幅

# 現在時刻を取得
datetime_now = datetime.now()
date_now = datetime_now.strftime('%Y/%m/%d %H:%M') # 例「2021/09/15 15:33」のような形式で取得
date_file = datetime_now.strftime('%Y%m%d')

# ディレクトリ設定
dir_path_data = os.path.join(os.getcwd(), 'data') # データ保存用
dir_temp_data = os.path.join(tempfile.gettempdir(), 'auto_detect_odp') # ダウンロードファイル保存用
if not os.path.isdir(dir_temp_data):
    os.makedirs(dir_temp_data)

#---------------------------------------#
# 関数設定（ファイル読み込み）
#---------------------------------------#
# 1. WebDriver起動
def open_driver():
    # Firefoxの設定
    profile = webdriver.FirefoxProfile()
    profile.set_preference("browser.download.folderList", 2) # 0：デスクトップ、1：既定、2：変更
    profile.set_preference("browser.download.dir", dir_temp_data) # 保存用ディレクトリ指定
    profile.set_preference("browser.download.manager.showWhenStarting", False) # ダウンロードマネージャー表示有無
    profile.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/pdf") # disk保存する形式は何にするか→PDF
    profile.set_preference("pdfjs.disabled", True) # ブラウザ上のPDFまたはJSのダウンロードの際、確認ダイアログを開かせない
    profile.set_preference("plugin.scan.plid.all", False) # プラグインを自動追加しない
    profile.set_preference("plugin.scan.Acrobat", "99.0") # PDFリーダーAcrobatのバージョン設定
    # browserの設定
    browser = webdriver.Firefox(profile)
    browser.implicitly_wait(setting['time_wait']) # 読み込みN秒待機
    browser_wait = WebDriverWait(browser, setting['time_wait']) # 指定の読み込み時間設定
    # ウィンドウ設定
    browser.set_window_size(window_w, window_h)
    # 出力
    return browser, browser_wait

# WebDriver終了
def close_driver(browser):
    browser.quit()

#---------------------------------------#
# 関数設定（ダウンロード実施判定）
#---------------------------------------#
# OAリスト
list_refusal = [
    # JP
#     '拒絶理由通知書', # 英文のみ対象のため、日本語文書はコメントアウトしている
    'Notice of Reasons for Refusal (TRANSLATED)',
#     '拒絶理由通知書（Notice of Reasons for Refusal）',
    # EP
    'European search opinion',
    # US
    'Non-Final Rejection',
    # KR
#     'Request for the Submission of an Opinion',
    'Request for the Submission of an Opinion (TRANSLATED)',
    # CN
    'Nth Office Action（Nth Office Action）',
    'First Office Action(PCT)',
]

# ダウンロード実施有無、ファイル名での選別
def check_download_file_name(country_name, file_name):
    # 確認
    check = 0
    check_length = len(file_name) # ファイル名の文字長
    # List内検索
    for letter in list_refusal:
        # 計算
        check_ratio = LS.ratio(letter, file_name) # LS距離の計算（letterとfile_name間）
        # 一致判定
        if letter in file_name or file_name in letter:  # 内部に一致する文字がある場合OK
            check = 1
            break
        elif check_ratio > settings['refusal_accuracy']:  # LS距離内での一致でOK
            check = 1
            print(check_ratio)
            break

    # 出力
    if check:
        return True
    else:
        return False

#---------------------------------------#
# アクセス用関数設定
#---------------------------------------#
# ブラウザをオープンしたあと、PatentScopeへ移動する
def access_ref(ref_title):
    print("◆%sにアクセスします。" % ref_title)
    browser.get('https://patentscope2.wipo.int/search/ja/search.jsf') # ブラウザでアクセス

# 特許No.を探す
def search_ref_while(search_id):
    browser.get('https://patentscope2.wipo.int/search/ja/search.jsf')
    # 種類選択用
    select_elem = browser.find_element_by_id('simpleSearchForm:field:input') # 目的のelementを取得（要確認：Google デベロッパーツール）
    select_elem.click() # 対象を物理クリック
    select_list = Select(select_elem) # クリックしたい対象を選択するために準備
    select_list.select_by_value('ALLNUM') # ALLNUMを選択する
    # No.を入力
    browser.find_element_by_id('simpleSearchForm:fpSearch:input').clear() # 検索の入力フォームをクリア
    browser.find_element_by_id('simpleSearchForm:fpSearch:input').send_keys(search_id) # 特許No.を入力
    # 検索開始
    browser.find_element_by_id('simpleSearchForm:fpSearch:buttons').click() # 検索ボタンをクリック

# search_idの設定（PatentScope用）
def exchange_id(search_id):
    search_id_split = search_id.split('-') # '-'で区切り、配列として出力
    search_id = "WO/%s/%s" % (search_id_split[0], search_id_split[1]) # %sの位置に代入する
    return search_id

# 検索試行
def search_ref(search_id):
    print("◆検索を行います。")
    # current_urlの読み込み
    search_ref_while(search_id) # 自作関数で特許No.を探す
    current_url = browser.current_url

    # 読み込みエラーを通り抜ける
    i = 0
    while current_url == 'https://patentscope2.wipo.int/search/ja/expired.jsf': # このurlのままだと読み込みエラーだと判断する
        print("検索に再度アクセスします。")
        search_ref_while(search_id_ch) # 自作関数で特許No.を探す
        current_url = browser.current_url # ブラウザの現行URLを取得
        i += 1
        if i > 10:
            print("データが読み込めません。このシステムのデータはなしとしてプログラムの実行を進めます。")
            break
    
    # bot 対策
    while current_url == 'https://patentscope2.wipo.int/search/ja/captcha/captcha.jsf': # ボット対策用URLを判定
        for i in range(2):
            beep(1000, 500)
        # ボット対策には、人手で対応することにした。画像認識判別も可能であるが、全く別の開発が必要になるため。
        print("Bot対応を行ってください。対応が終わったら、何らかの文字列（例：1）を入力してOKボタンを押してください。")
        input_bot = input()
        while input_bot is None: # Inputされない場合は待機
            time.sleep(30)
        # 更新
        current_url = browser.current_url

# 検索画面のデータ確認
def check_search_data():
    check_elem = browser.find_elements_by_class_name('patent-family-member') # 検索結果後の表を取得
    if len(check_elem) > 0:
        print("◆各国データが存在します。読み込み中です。")
        return True
    else:
        print("◆データが存在しません。WOのみで抽出を行います。")
        return False

# urlを取得
def read_search_data(search_id):
    # データ確認
    dict_link_url = {search_id: ''}
    check_elem = browser.find_elements_by_class_name('patent-family-member') # 各国へのリンクリストを取得
    if len(check_elem) > 0:
        # 各国URLを取得
        for i in range(0, len(check_elem)): # 繰り返し設定
            links = check_elem[i].find_elements(By.TAG_NAME, "a") # aタグを見つける
            for j in range(0, len(links)):
                co_text = links[j].text # リンクテキスト（国名）を取得
                co_link = links[j].get_attribute("href") # hrefはリンク内容のためそれを取得
                dict_link_url[co_text] = co_link # key, valueを国名, リンクで設定
    
    return dict_link_url

# 各国のデータを取得する
def read_reach_data(df_dict, list_data_columns, dict_link_url):
    # 各国の書類にアクセスする
    for key, key_url in dict_link_url.items():
        # key_head読み込み
        key_head = re.findall('[a-z]+', key, flags=re.IGNORECASE) # 正規表現で国名を取得（異常な文字が含まれる場合、読み込まない）
        key_head_name = key_head[0]
        print("〇", key, "を読み込み中です...")
        
        # アクセス（WOは意図的に読み込まない）
        if len(key_url) > 0:
            browser.get(key_url)
            
        # チェック
        check_elem = browser.find_elements_by_link_text('書類') # 「書類」というリンクがあれば読み込みを行う
        if len(check_elem) > 0:
            print("書類がありました。")
            # 別関数での読み込み実施
            for count in range(setting['count_retry']):
                try:
                    # 書類からの呼び出し
                    df_dict = read_reach_data_documents(df_dict, key_head_name) # 別関数で指定
                    break
                except exceptions.StaleElementReferenceException as e: # エラー回避
                    time.sleep(setting['time_wait_q'])
        else:
            print("書類が存在しないか、読み込み不良のため読み込みません。")

    # 出力
    return df_dict

# 個別の書類ごとに読み込む
def read_reach_data_documents(df_dict, key_head_name):
    # '書類'クリック
    if expected_conditions.visibility_of_all_elements_located((By.CLASS_NAME, 'b-infobox--has-spinner'))    or expected_conditions.visibility_of_all_elements_located((By.CLASS_NAME, 'ui-blockui')): # クリック要素上に別要素が存在してしまっている場合待機
        time.sleep(setting['time_wait_q'])
    browser_wait.until(expected_conditions.element_to_be_clickable((By.LINK_TEXT, '書類'))) # 「書類」がクリックできるまで待機
    select_elem = browser.find_element_by_link_text('書類')
    select_elem.click()
    
    # 読み込みwait
    time.sleep(setting['time_wait_q'])
    # テーブル存在確認
    patents_elem = browser_wait.until(expected_conditions.visibility_of_all_elements_located((By.CLASS_NAME, 'ui-datatable')))
    if len(patents_elem) > 0:
        # DataFrame初期設定
        if key_head_name not in df_dict:
            df_dict[key_head_name] = pd.DataFrame(columns=list_data_columns)

        # テーブルごとに取得
        for i in range(len(patents_elem)):
            # 読み込み
            table_header = patents_elem[i].find_elements(By.CLASS_NAME, "ui-datatable-header")
            if len(table_header) > 0:
                if table_header[0].text == "公開された出願":
                    print("No.%sは公開された出願のため、データを取得しません。" % i)
                    # table_header[0].textをdfに入れておきたい
                else:
                    print("No.%sを取得します。" % i)
                    # 表データを全て読み込み可能な状態になるまで待機
                    browser_wait.until(expected_conditions.visibility_of_all_elements_located((By.CLASS_NAME, "ui-datatable-tablewrapper")))
                    # 読み込み
                    table_body = patents_elem[i].find_elements(By.CLASS_NAME, "ui-datatable-tablewrapper")
                    if len(table_body) > 0:
                        # list_allセット
                        list_all = []
                        trs = table_body[0].find_elements(By.TAG_NAME, "tr")
                        # 表の先頭行（tr）を含まず取得
                        for j in range(1, len(trs)):
                            # リストセット
                            list = []
                            # 表のセル（td）ごとに取得
                            exp_tds = trs[j].find_elements(By.TAG_NAME, "td")
                            # tds
                            if len(exp_tds) > 2:
                                ## df取得
                                # 日付抽出（元データ例：21.08.2010）
                                tds0_text = exp_tds[0].text.split('.')
                                exp_date = tds0_text[2]+"-"+tds0_text[1]+"-"+tds0_text[0]
                                # FileName
                                exp_file_name = exp_tds[1].text
                                # Key
                                exp_key = ref_title +'_'+ key_head_name +'_'+ exp_date.replace('-', '') +'_'+ exp_file_name

                                # Date
                                list.append(exp_date)
                                # Reference
                                list.append(ref_title)
                                # FileName
                                list.append(exp_file_name)
                                # Key
                                list.append(exp_key)
                                # リンクを取得
                                exp_hrefs = trs[j].find_elements_by_link_text('PDF')
                                if len(exp_hrefs) > 0:
                                    list.append(exp_hrefs[0].get_attribute("href"))
                                else:
                                    # PDFリンクが存在しない場合
                                    exp_hrefs = trs[j].find_elements(By.TAG_NAME, "a")
                                    if len(exp_hrefs) > 0:
                                        list.append(exp_hrefs[0].get_attribute("href"))
                                    else:
                                        list.append("")
                                # list_allに追加
                                list_all.append(list)
                                
                                ## ダウンロード実施
                                # ファイル名判定機能（別関数）
                                if check_download_file_name(key_head_name, exp_file_name) and len(exp_hrefs) > 0:
                                    print("ダウンロード実施【%s】%s" % (key_head_name, exp_file_name))
                                    # ダウンロードフォルダ内を全削除する
                                    shutil.rmtree(dir_temp_data)
                                    # ダウンロードリンクの指定
                                    tag_download = exp_hrefs[0]
                                    # リンクまでスクロール
                                    tag_download.location_once_scrolled_into_view
                                    # 新しいウィンドウで開いてダウンロード（Ctrlを押す⇒リンクタグをクリック⇒Ctrlを離す）
                                    ActionChains(browser).key_down(Keys.CONTROL).click(tag_download).key_up(Keys.CONTROL).perform()
                                    # ダウンロードしている間はwait
                                    time.sleep(setting['time_wait_file'])
                                    # ファイル名を変更する
                                    exp_file_fname = exp_file_name.replace(' ', '_')+'.pdf'
                                    exp_file_path = os.path.join(dir_temp_data, exp_file_fname)
                                    exp_key_fname = exp_key+'.pdf'
                                    exp_key_path = os.path.join(dir_odp, exp_key_fname)
                                    # ファイルが存在する場合
                                    # print(exp_file_path)
                                    if os.path.isfile(exp_file_path):
                                        print("ファイルをダウンロードしました")
                                        if os.path.isfile(exp_key_path) is False:
                                            os.rename(exp_file_path, exp_key_path)
                                    else:
                                        print("ダウンロードできませんでした")
                                        dict_download[exp_key_fname] = tag_download.get_attribute("href") # ダウンロード出来なかった場合は別途データ保管・表示して、作業者がダウンロードできるようにする
                                
                        # dfに追加
                        df_dict[key_head_name] = pd.concat([df_dict[key_head_name], pd.DataFrame(list_all, columns=list_data_columns)], axis=0)
                        
    # 出力
    return df_dict

#---------------------------------------#
# 1. プログラム実行
#---------------------------------------#
# 特許No.入力
print("WO番号を「XXXX-XXXXXX」（例：2016-031178）の形式で入力してください。")
search_id = input()

# dir設定
dir_odp = os.path.join(dir_path_data, search_id)
if os.path.isdir(dir_odp) is False:
    os.mkdir(dir_odp)
    print("Dataフォルダ「%s」を作成しました" % dir_odp)

# J-PlatPatの読み込み
# https://www.j-platpat.inpit.go.jp/
# 今回は対応していません。（ページの変更が多すぎるため）

# PatentScopeの読み込み
# https://patentscope2.wipo.int/search/ja/search.jsf
ref_title = 'PatentScope' # 参考サイトタイトル

# start_driver
browser, browser_wait = open_driver()
# アクセス
access_ref(ref_title)
# id変更
seach_id_PS = exchange_id(search_id) # PatentScope用id
# 検索
search_ref(seach_id_PS)
# 検索結果チェック
if check_search_data():
    dict_link_url = read_search_data(seach_id_PS)
    # データ読み込み
    if len(dict_link_url) > 0:
        df_dict = read_reach_data(df_dict, list_data_columns, dict_link_url)
# close_driver
close_driver(browser)

#---------------------------------------#
# 2. エクセル出力（基本情報）
#---------------------------------------#
# データ後処理
print("出力のためにデータの整形を行います。")

# df_dictの最新順化
for key, key_dict in df_dict.items():
    print(key, "を処理中です...")
    # Date処理
    key_dict['Date'] = pd.to_datetime(key_dict.Date)
    key_dict.sort_values(['Date', 'FileName', 'Reference'], ascending=False, inplace=True)
    
    # 最新状況取得
    if key not in df_all.index:
        df_all.loc[key] = ""
        df_all.at[key, 'C.Code'] = key
    if len(key_dict) > 0:
        df_all.at[key, '最新情報'] = key_dict.iloc[0]['FileName']
        df_all.at[key, '最終更新日'] = key_dict.iloc[0]['Date']

# エクセルを一時作成（表示形式等は作成してからでないと変更できないため）
excel_name = os.path.join(dir_path_data, date_file + '_WO' + search_id + '.xlsx')
print("エクセルを" + excel_name + "で出力中です...")
df_all.to_excel(excel_name, sheet_name='基礎情報', index=False)

# エクセルの再編集
wb = openpyxl.load_workbook(excel_name)
ws = wb['基礎情報']

# 日時表示形式変更
i = 2
for date in df_all['最終更新日']:
    ws.cell(row=i, column=3).number_format = "yyyy/mm/dd"
    i += 1

# セル幅調整
for i, col in enumerate(ws.columns):
    max_length = 0
    column = col[0].column

    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))

    # 日付の大きさを調整（要調整）
    if i == 2:
        max_length = 10
    elif i == 1:
        max_length = 70
    
    adjusted_width = (max_length + 1) * 1.05
    ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

# 保存
wb.save(excel_name)

#---------------------------------------#
# 3. 各国最新情報取得
#---------------------------------------#
sp_dict = {}
for key, key_dict in df_dict.items():
    print(key, "を処理中です...")
    # df用意
    copy_dict = key_dict.copy()
    copy_dict.loc[(copy_dict["Url"] != ""), "Url"] = "リンク"
    copy_dict.drop(columns=['Key'], inplace=True)

    # エクセルに書き込み
    with pd.ExcelWriter(excel_name) as xls_writer:
        xls_writer.book = openpyxl.load_workbook(excel_name)
        copy_dict.to_excel(xls_writer, sheet_name=key, index=False)

    # 再読み込み（表示形式は作成してからでないと変更できない）
    wb = openpyxl.load_workbook(excel_name)
    ws = wb[key]

    # 日時表示形式変更
    i = 2
    for date in key_dict['Date']:
        ws.cell(row=i, column=1).number_format = "yyyy/mm/dd"
        i += 1

    # リンク処理
    i = 2
    check_i = 0
    for link_url in key_dict['Url']:
        # Urlは3列目
        if link_url:
            ws.cell(row=i, column=4).hyperlink = (link_url)
            check_i += 1
        i += 1
    print("リンクを" + str(check_i) + "件追加しました")

    # セル幅調整
    for i, col in enumerate(ws.columns):
        max_length = 0
        column = col[0].column

        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        # 日付要調整
        if i == 0:
            max_length = 10

        adjusted_width = (max_length + 1) * 1.05
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

    # 保存
    wb.save(excel_name)

# 完了
print("エクセルの作成が終了しました。")
beep(1000, 1000)

#---------------------------------------#
# 4. その他出力
#---------------------------------------#
if len(dict_download):
    print("★何らかの原因でダウンロードできなかったファイルは以下の通りです。\n手動ダウンロードを行ってください。")
    print(dict_download)
